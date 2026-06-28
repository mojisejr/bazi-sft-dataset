# -*- coding: utf-8 -*-
"""
Distill the 3 Excel knowledge files in knownlage/ปฏิกิริยาธาตุ/ into JSON assets
consumed by the pair-matching engine (src/lib/bazi/pair-matching.ts).

Outputs (src/lib/bazi/data/pair/):
  pair-matrix.json   - { "<domain>": { "<ourStem><ourBranch>|<pStem><pBranch>": {...} } }
  rating-scale.json  - { "<domain>": [ {min,max,grade,emoji,text} ] }
  sising.json        - 12 deity stars (B1..B12) descriptive knowledge
  reference.json     - nisai (day-pillar personality), shengxia role/love readings, kubun raw

Run:  python scripts/distill-pair-knowledge.py
"""
import json
import os
import unicodedata

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC_DIR = os.path.join(ROOT, "knownlage", "ปฏิกิริยาธาตุ")
OUT_DIR = os.path.join(ROOT, "src", "lib", "bazi", "data", "pair")

STEMS = set("甲乙丙丁戊己庚辛壬癸")
BRANCHES = set("子丑寅卯辰巳午未申酉戌亥")

WORK_FILE = "คู่สมพงษ์(การงาน).xlsx"
LOVE_FILE = "คู่สมพงษ์(ความรัก).xlsx"
SISING_FILE = "12 สี่ซิ้ง.xlsx"


def norm(value):
    """Normalise a cell to a trimmed, NFKC-canonical string (fixes U+F971 -> 辰 etc)."""
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load(fname):
    return openpyxl.load_workbook(os.path.join(SRC_DIR, fname), data_only=True)


# ---------------------------------------------------------------------------
# 1. pair-matrix.json
# ---------------------------------------------------------------------------
def parse_pair_sheets(wb, domain):
    """Each pair sheet = one of-our-pillars x ~60 partner-pillars, in 2-row blocks.
    Top row:    B=ourStem D=partnerStem E=stemCode F=branchCode G=sisingCode H=ratio I=points J,K,L=components
    Bottom row: B=ourBranch D=partnerBranch  G=sisingName
    percent = mean(J,K,L). grade derived later from rating buckets.
    stemCode (E) = relation ของก้านเรา×ก้านเขา · branchCode (F) = ของกิ่งเรา×กิ่งเขา
    (โค้ด A1..A12 = 12 เชี่ยงแซ; ใช้คู่กับ reference role lists เพื่อหาคำทำนายรายแท่ง).
    """
    out = {}
    count = 0
    for ws in wb.worksheets:
        title = ws.title.strip()
        # pair sheets are named like "ม1,1" / "ฟ2,4" - contain a comma + element prefix
        if "," not in title or title[0] not in "มฟดทน":
            continue
        rows = list(ws.iter_rows(values_only=True))

        def cell(r, idx):
            if r < 0 or r >= len(rows):
                return ""
            row = rows[r]
            return norm(row[idx]) if idx < len(row) else ""

        i = 0
        while i < len(rows):
            our_stem = cell(i, 1)       # B
            partner_stem = cell(i, 3)   # D
            if our_stem in STEMS and partner_stem in STEMS:
                our_branch = cell(i + 1, 1)
                partner_branch = cell(i + 1, 3)
                if our_branch in BRANCHES and partner_branch in BRANCHES:
                    row = rows[i]
                    comps = [num(row[9]), num(row[10]), num(row[11])]  # J,K,L
                    comps = [c for c in comps if c is not None]
                    percent = round(sum(comps) / len(comps), 2) if comps else None
                    stem_code = cell(i, 4)         # E (โค้ดก้านเรา×ก้านเขา)
                    branch_code = cell(i, 5)       # F (โค้ดกิ่งเรา×กิ่งเขา)
                    sising_code = cell(i, 6)        # G top
                    sising_name = cell(i + 1, 6)    # G bottom
                    key = f"{our_stem}{our_branch}|{partner_stem}{partner_branch}"
                    out[key] = {
                        "domain": domain,
                        "percent": percent,
                        "components": comps,
                        "points": num(row[8]),       # I
                        "ratio": num(row[7]),        # H
                        "stemCode": stem_code or None,
                        "branchCode": branch_code or None,
                        "sisingCode": sising_code or None,
                        "sisingName": sising_name or None,
                    }
                    count += 1
                    i += 2
                    continue
            i += 1
    return out, count


# ---------------------------------------------------------------------------
# 2. rating-scale.json
# ---------------------------------------------------------------------------
def parse_rating(wb, sheet_name):
    """Rows: level(1-10), min%, (maybe '-'), max%, emoji, text. Returns ordered buckets.
    The leading level/star-count integer (1..10) is dropped; the two remaining numbers
    are the percentage bounds (no real bound falls in 1..10 except the 0.0 minimum)."""
    ws = wb[sheet_name]
    buckets = []
    for row in ws.iter_rows(values_only=True):
        nums, emoji, text = [], None, None
        for v in row:
            s = norm(v)
            if s == "" or s == "-":
                continue
            n = num(v)
            if n is not None:
                nums.append(n)
            elif any(ch in s for ch in "💸♥️"):
                emoji = s
            elif len(s) > 8:
                text = s
        # drop the first level integer (1..10) so only the % bounds remain
        for idx, n in enumerate(nums):
            if n.is_integer() and 1 <= n <= 10:
                nums.pop(idx)
                break
        pcts = [n for n in nums if 0 <= n <= 100]
        if len(pcts) >= 2 and text:
            lo, hi = min(pcts), max(pcts)
            buckets.append({"min": round(lo, 2), "max": round(hi, 2),
                            "emoji": emoji, "text": text})
    buckets.sort(key=lambda b: b["min"])
    return buckets


def parse_grade_scale(wb):
    """The 13-step letter grade legend (A+..F) lives in any pair sheet, cols Q,R,S
    (indices 16,17,18) rows 1..13: min%, max%, gradeLetter."""
    for ws in wb.worksheets:
        title = ws.title.strip()
        if "," not in title or title[0] not in "มฟดทน":
            continue
        grades = []
        for row in ws.iter_rows(min_row=1, max_row=14, values_only=True):
            lo = num(row[16]) if len(row) > 16 else None
            hi = num(row[17]) if len(row) > 17 else None
            grade = norm(row[18]) if len(row) > 18 else ""
            if lo is not None and hi is not None and grade:
                grades.append({"min": round(lo, 2), "max": round(hi, 2), "grade": grade})
        if grades:
            grades.sort(key=lambda g: g["min"])
            return grades
    return []


# ---------------------------------------------------------------------------
# 3. sising.json  (12 deity stars, codes B1..B12)
# ---------------------------------------------------------------------------
def parse_sising(wb):
    ws = wb["12 สี่ซิ้ง"]
    rows = list(ws.iter_rows(values_only=True))
    header = [norm(c) for c in rows[0]]
    out = []
    branch_cols = list(range(2, 14))  # C..N = 子..亥
    for row in rows[1:]:
        code = norm(row[0])
        if not code:
            continue
        positions = [norm(row[c]) for c in branch_cols]
        out.append({
            "code": code,
            "score": num(row[1]),
            "branchPositions": positions,
            "nameCn": norm(row[14]),
            "nameTh": norm(row[15]),
            "short": norm(row[16]),
            "long": norm(row[17]),
            "aspects": {
                "work": norm(row[18]),
                "money": norm(row[19]),
                "love": norm(row[20]),
                "family": norm(row[21]),
                "business": norm(row[22]),
                "health": norm(row[23]),
            },
            "summary": norm(row[24]),
        })
    return out


# ---------------------------------------------------------------------------
# 4. reference.json
# ---------------------------------------------------------------------------
def parse_nisai(wb):
    """นิสัยหลักวันบนล่าง: 3-row blocks (stem / branch / element+stage) each with text in col D."""
    ws = wb["นิสัยหลักวันบนล่าง"]
    rows = list(ws.iter_rows(values_only=True))
    by_stem, by_branch, by_stage = {}, {}, {}
    for r in rows:
        a, b, d = norm(r[0]), norm(r[1]) if len(r) > 1 else "", norm(r[3]) if len(r) > 3 else ""
        if not d:
            continue
        if b in STEMS:
            by_stem[b] = d
        elif b in BRANCHES:
            by_branch[b] = d
        elif b:  # stage name e.g. หมกยก
            by_stage[b] = d
    return {"byStem": by_stem, "byBranch": by_branch, "byStage": by_stage}


def parse_role(wb, sheet_name, code_col, name_col, score_col, text_col):
    """Role / 12-shengxia sheets: header row maps stem -> column; each data row is a
    shengxia stage with {code,name,score,narrative} plus the branch it occupies per stem.
    Returns { 'stages':[...], 'stemColumns': {stem: colIndex} }."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    stem_cols = {}
    header_idx = None
    for ri, r in enumerate(rows):
        for ci, v in enumerate(r):
            s = norm(v)
            if s in STEMS:
                stem_cols[s] = ci
        if len(stem_cols) >= 8:
            header_idx = ri
            break
    stages = []
    for r in rows[(header_idx or 0) + 1:]:
        code = norm(r[code_col]) if len(r) > code_col else ""
        name = norm(r[name_col]) if len(r) > name_col else ""
        text = norm(r[text_col]) if len(r) > text_col else ""
        if not code or not text:
            continue
        branch_by_stem = {}
        for stem, ci in stem_cols.items():
            b = norm(r[ci]) if len(r) > ci else ""
            if b in BRANCHES:
                branch_by_stem[stem] = b
        stages.append({
            "code": code,
            "name": name,
            "score": num(r[score_col]) if len(r) > score_col else None,
            "narrative": text,
            "branchByStem": branch_by_stem,
        })
    return stages


def parse_kubun_raw(wb):
    ws = wb["คู่บุญ คู่กรรม"]
    out = []
    for r in ws.iter_rows(values_only=True):
        cells = [norm(c) for c in r]
        if any(cells):
            out.append(cells)
    return out


# ---------------------------------------------------------------------------
def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    work_wb = load(WORK_FILE)
    love_wb = load(LOVE_FILE)
    sising_wb = load(SISING_FILE)

    # pair matrix
    work_matrix, wc = parse_pair_sheets(work_wb, "work")
    love_matrix, lc = parse_pair_sheets(love_wb, "love")
    matrix = {"work": work_matrix, "love": love_matrix}
    write_json("pair-matrix.json", matrix)
    print(f"pair-matrix: work={wc} love={lc} combos")

    # rating
    rating = {
        "grades": parse_grade_scale(work_wb),
        "work": parse_rating(work_wb, "เรตติ้งหุ้นส่วน"),
        "love": parse_rating(love_wb, "เรตติ้งคะแนน"),
    }
    write_json("rating-scale.json", rating)
    print(f"rating: grades={len(rating['grades'])} "
          f"work={len(rating['work'])} love={len(rating['love'])} buckets")

    # sising
    sising = parse_sising(sising_wb)
    write_json("sising.json", sising)
    print(f"sising: {len(sising)} stars")

    # reference
    reference = {
        "nisai": parse_nisai(work_wb),
        "roleBoss": parse_role(work_wb, "คำทำนายตัวเรา>เจ้านาย", 1, 2, 3, 4),
        "roleSubordinate": parse_role(work_wb, "คำทำนาย ลูกน้อง>ตัวเรา", 1, 2, 3, 4),
        "rolePartner": parse_role(work_wb, "คำทำนายหุ้นส่วนเพื่อนร่วมงาน", 1, 2, 3, 4),
        "loveShengxia": parse_role(love_wb, "12เชี่ยงแซความรัก", 0, 1, 2, 3),
        "kubunRaw": parse_kubun_raw(work_wb),
    }
    write_json("reference.json", reference)
    print("reference: "
          f"nisai stems={len(reference['nisai']['byStem'])} "
          f"boss={len(reference['roleBoss'])} sub={len(reference['roleSubordinate'])} "
          f"partner={len(reference['rolePartner'])} love12={len(reference['loveShengxia'])}")


def write_json(name, data):
    path = os.path.join(OUT_DIR, name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
