# -*- coding: utf-8 -*-
"""Build the repo-local distilled corpus (knownlage/distilled/) from docs/ originals.

hybrid-retrieval (src/lib/bazi/hybrid-retrieval.ts) reads each dimension's
sourceRelativePaths from the external all_distilled corpus, and falls back to
knownlage/distilled/<relativePath> when the external corpus is absent. The
external corpus is not shipped, so we mirror every referenced source here by
converting the matching docx -> .md and xlsx sheet -> .csv.

The TARGETS list is the union of:
  - dictionary sourceRelativePaths (src/lib/bazi/dictionaries/*.ts)
  - registry TIER_B_SOURCE_PATHS (src/lib/bazi/hybrid-retrieval-registry.ts)

Run:  python scripts/build-distilled-corpus.py
"""
import csv
import io
import os
import sys
import unicodedata

import docx
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCS = os.path.join(ROOT, "docs")
OUT = os.path.join(ROOT, "knownlage", "distilled")


def norm(text):
    """NFKC + recombine สระอำ ที่สะกดเพี้ยน (U+0E4D+U+0E32 -> U+0E33) และ 辰 U+F971 -> U+8FB0."""
    if text is None:
        return ""
    return unicodedata.normalize("NFKC", str(text)).replace("ํา", "ำ")


# (target relativePath under knownlage/distilled, kind, source basename, sheet|None)
TARGETS = [
    # ----- docx -> md -----
    ("1.นิสัยโดยพื้นฐาน/1.นิสัยโดยพื้นฐาน.md", "docx", "1.นิสัยโดยพื้นฐาน.docx", None),
    ("Source3_ สุขภาพ(พื้นฐาน)/Source3_ สุขภาพ(พื้นฐาน).md", "docx", "Source3_ สุขภาพ(พื้นฐาน).docx", None),
    ("Source4_ การเงินและการลงทุน/Source4_ การเงินและการลงทุน.md", "docx", "Source4_ การเงินและการลงทุน.docx", None),
    ("Source5_ ความรักและความสัมพันธ์/Source5_ ความรักและความสัมพันธ์.md", "docx", "Source5_ ความรักและความสัมพันธ์.docx", None),
    ("Source6_ การงานและธุรกิจ/Source6_ การงานและธุรกิจ.md", "docx", "Source6_ การงานและธุรกิจ.docx", None),
    ("การงานและธุรกิจ/การงานและธุรกิจ.md", "docx", "การงานและธุรกิจ.docx", None),
    ("การทายวัยจร/การทายวัยจร.md", "docx", "การทายวัยจร.docx", None),
    ("การเงินและการลงทุน/การเงินและการลงทุน.md", "docx", "การเงินและการลงทุน.docx", None),
    ("ความรักและความสัมพันธ์/ความรักและความสัมพันธ์.md", "docx", "ความรักและความสัมพันธ์.docx", None),
    ("ชงเฮ้งไห่ผั่วภาคี(เนื้อหา).docx.md", "docx", "ชงเฮ้งไห่ผั่วภาคี(เนื้อหา).docx", None),
    ("ตาราง 12 เชี่ยงแซ/ตาราง 12 เชี่ยงแซ.md", "docx", "ตาราง 12 เชี่ยงแซ.docx", None),
    ("ตารางชงเฮ้งไห่ผั่ว/ตารางชงเฮ้งไห่ผั่ว.md", "docx", "ตารางชงเฮ้งไห่ผั่ว.docx", None),
    ("ระบบ 12 เชี่ยงแซ 十二長生.md", "docx", "ระบบ 12 เชี่ยงแซ 十二長生.docx", None),
    ("สุขภาพ(พื้นฐาน)/สุขภาพ(พื้นฐาน).md", "docx", "สุขภาพ(พื้นฐาน).docx", None),
    ("Step การอ่านดวง/Step การอ่านดวง.md", "docx", "Step การอ่านดวง.docx", None),
    ("Source1_ Step การอ่านดวง/Source1_ Step การอ่านดวง.md", "docx", "Source1_ Step การอ่านดวง.docx", None),
    ("เกณฑ์ความแข็งอ่อน_ดวง5แบบ/2026-04-23_strength-evaluation-step.md", "docx",
     "Step พิจารณาดวงแข็งเกินไป แข็งแรง สมุดล อ่อนแอ อ่อนแอเกินไป.docx", None),
    ("ตารางปฏิกิริยาธาตุ/ตารางปฏิกิริยาธาตุ.md", "docx", "ตารางปฏิกิริยาธาตุ.docx", None),
    ("อธิบายวงจรธาตุ/อธิบายวงจรธาตุ.md", "docx", "อธิบายวงจรธาตุ.docx", None),
    ("12สี่ซิ้ง/12สี่ซิ้ง.md", "docx", "12สี่ซิ้ง.docx", None),
    ("ตำราโหราศาสตร์เคี้ยงคุง/ตำราโหราศาสตร์เคี้ยงคุง.md", "docx", "ตำราโหราศาสตร์เคี้ยงคุง.docx", None),
    ("Stepพิจารณาดวงแข็งเกินไป-แข็งแรง-อ่อนแอ-อ่อนแอเกินไป.md", "docx",
     "Step พิจารณาดวงแข็งเกินไป แข็งแรง สมุดล อ่อนแอ อ่อนแอเกินไป.docx", None),
    ("ตำรา24สารท/ตำรา24สารท.md", "docx", "ตำรา24สารท.docx", None),
    ("Source7_ การเสริมดวง/Source7_ การเสริมดวง.md", "docx", "Source7_ การเสริมดวง.docx", None),
    # ----- xlsx sheet -> csv -----
    ("คู่สมพงษ์(ความรัก)/คู่สมพงษ์(ความรัก) - หลักวันเท่านั้น.csv", "csv",
     "คู่สมพงษ์(ความรัก).xlsx", "หลักวันเท่านั้น"),
    ("ลักษณะ ของ ดิถี 10 ราศีบน 60กะจื่อ วัยจร/ลักษณะ ของ ดิถี 10 ราศีบน 60กะจื่อ วัยจร - ข้อมูลช่องนิสัย.csv", "csv",
     "ลักษณะ ของ ดิถี 10 ราศีบน 60กะจื่อ วัยจร.xlsx", "ข้อมูลช่องนิสัย"),
    ("ลักษณะ ของ ดิถี 10 ราศีบน 60กะจื่อ วัยจร/ลักษณะ ของ ดิถี 10 ราศีบน 60กะจื่อ วัยจร - อาชีพถูกดวง.csv", "csv",
     "ลักษณะ ของ ดิถี 10 ราศีบน 60กะจื่อ วัยจร.xlsx", "อาชีพถูกดวง"),
    ("ลักษณะนิสัย60แบบ_ราศีบน-ล่าง_12เซี่ยงแซ/ลักษณะนิสัย60แบบ_ราศีบน-ล่าง_12เซี่ยงแซ - นิสัยราศีบน,ล่าง,เซี่ยงแซ.csv", "csv",
     "ลักษณะนิสัย60แบบ_ราศีบน-ล่าง_12เซี่ยงแซ.xlsx", "นิสัยราศีบน,ล่าง,เซี่ยงแซ"),
    ("สูตรคำนวณวัยจรลัคนา/2026-04-23_lagna-formula.csv", "csv", "คำนวนวัยจร ลัคนา.xlsx", "คำนวนลัคนา"),
    ("สูตรคำนวณวัยจรลัคนา/2026-04-23_major-luck-formula.csv", "csv", "คำนวนวัยจร.xlsx", "วัยจรใหญ่"),
]


def find_source(basename):
    """Resolve a docx/xlsx basename under docs/, preferring the Mootech AI root over dated/variant subdirs."""
    matches = []
    for root, _, files in os.walk(DOCS):
        for f in files:
            if f == basename and not f.startswith("~$"):
                matches.append(os.path.join(root, f))
    if not matches:
        return None

    def rank(p):
        rel = os.path.relpath(p, DOCS)
        # prefer plain "Mootech AI/<name>" or "docs/<name>"; deprioritize version dirs
        bad = any(seg in rel for seg in ("25.4", "28.4", os.sep + "new" + os.sep, "Source 2"))
        depth = rel.count(os.sep)
        return (bad, depth, len(rel))

    return sorted(matches, key=rank)[0]


def iter_block_items(doc):
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph
    for child in doc.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, doc)
        elif isinstance(child, CT_Tbl):
            yield Table(child, doc)


def table_to_md(table):
    lines = []
    for r, row in enumerate(table.rows):
        cells = [norm(c.text).replace("\n", " ").replace("|", "/").strip() for c in row.cells]
        lines.append("| " + " | ".join(cells) + " |")
        if r == 0:
            lines.append("| " + " | ".join(["---"] * len(cells)) + " |")
    return "\n".join(lines)


def docx_to_md(path):
    doc = docx.Document(path)
    out = []
    for block in iter_block_items(doc):
        if block.__class__.__name__ == "Paragraph":
            t = norm(block.text).strip()
            if t:
                out.append(t)
        else:
            md = table_to_md(block)
            if md.strip():
                out.extend(["", md, ""])
    return "\n".join(out).strip() + "\n"


def xlsx_sheet_to_csv(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"sheet {sheet!r} not in {wb.sheetnames}")
    ws = wb[sheet]
    buf = io.StringIO()
    w = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if c is None else norm(c).strip() for c in row])
    return buf.getvalue()


def main():
    total = len(TARGETS)
    wrote = 0
    skipped = []
    for i, (rel, kind, src_basename, sheet) in enumerate(TARGETS, 1):
        src = find_source(src_basename)
        if not src:
            skipped.append((rel, f"source not found: {src_basename}"))
            print(f"[{i}/{total}] SKIP  {rel}  (no source: {src_basename})")
            continue
        try:
            content = docx_to_md(src) if kind == "docx" else xlsx_sheet_to_csv(src, sheet)
        except Exception as e:  # noqa: BLE001 - report and continue
            skipped.append((rel, str(e)))
            print(f"[{i}/{total}] SKIP  {rel}  ({e})")
            continue
        dst = os.path.join(OUT, rel)
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        with open(dst, "w", encoding="utf-8") as fh:
            fh.write(content)
        wrote += 1
        print(f"[{i}/{total}] wrote {rel}  ({len(content)} chars)")

    print(f"\n=== DONE: {wrote}/{total} written, {len(skipped)} skipped ===")
    for rel, why in skipped:
        print(f"  SKIPPED {rel}: {why}")
    if skipped:
        sys.exit(0)  # non-fatal; report only


if __name__ == "__main__":
    main()
