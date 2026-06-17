#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
สกัดขอบสารท (ปฏิทิน 150 ปี) -> src/lib/bazi/data/almanac/solar-terms-2450-2600.json

ที่มา: knownlage/2450-2600วันเปลี่ยนสารทเล็กสารทใหญ่.xlsx
โครงคอลัมน์ (ยืนยันจาก SQL ในคอลัมน์ 35):
  1=ค.ศ. 2=พ.ศ. 3,4=เสาปี(干支) 5=เดือน(1=ม.ค.) 7,8=เสาเดือน(干支)
  9,10,11  = start (วันเริ่มต้นเดือน) = 節 (สารทใหญ่/วันเปลี่ยนเดือน)
  15,16,17 = big_start (วันเริ่มต้นสารทใหญ่) — เท่ากับ start
  21,22,23 = small_start (วันเริ่มต้นสารทเล็ก) = 中氣

ต่างจาก extract-almanac-tables.py ตรงที่ "เพิ่ม small_start_*" ที่เดิมถูกตัดทิ้ง
รัน:  PYTHONUTF8=1 python scripts/extract-solar-terms.py
"""
import json, os, unicodedata

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "knownlage", "2450-2600วันเปลี่ยนสารทเล็กสารทใหญ่.xlsx")
OUT = os.path.join(ROOT, "src", "lib", "bazi", "data", "almanac", "solar-terms-2450-2600.json")


def _norm_cjk(s):
    # normalize เฉพาะ CJK compatibility ideographs (U+F900-FAFF เช่น 辰=U+F971)
    # ห้าม NFKC ทั้งสตริง เพราะจะทำลายสระไทย (ตรงกับ extract-almanac-tables.py)
    return "".join(unicodedata.normalize("NFKC", ch) if "豈" <= ch <= "﫿" else ch for ch in s)


def v(ws, r, c):
    val = ws.cell(r, c).value
    if val is None:
        return None
    if hasattr(val, "strftime"):  # time cell
        return val.strftime("%H:%M")
    return _norm_cjk(str(val)).strip()


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    out = {}
    for r in range(2, ws.max_row + 1):
        be, mon = ws.cell(r, 2).value, ws.cell(r, 5).value
        if be is None or mon is None:
            continue
        try:
            be, mon = int(be), int(mon)
        except (TypeError, ValueError):
            continue
        ystem, ybr = v(ws, r, 3), v(ws, r, 4)
        mstem, mbr = v(ws, r, 7), v(ws, r, 8)
        out.setdefault(str(be), {})[str(mon)] = {
            "year_pillar": (ystem or "") + (ybr or ""),
            "month_pillar": (mstem or "") + (mbr or ""),
            "start_day": v(ws, r, 9), "start_month": v(ws, r, 10), "start_time": v(ws, r, 11),
            "big_start_day": v(ws, r, 15), "big_start_month": v(ws, r, 16), "big_start_time": v(ws, r, 17),
            "small_start_day": v(ws, r, 21), "small_start_month": v(ws, r, 22), "small_start_time": v(ws, r, 23),
        }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    yrs = sorted(out, key=int)
    print(f"wrote {os.path.relpath(OUT, ROOT)}: {len(out)} years (range {yrs[0]}-{yrs[-1]})")
    sample = out[yrs[0]]["1"]
    print("sample 2450/1:", json.dumps(sample, ensure_ascii=False))


if __name__ == "__main__":
    main()
