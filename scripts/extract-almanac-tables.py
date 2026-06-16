#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
แกะตาราง lookup จากไฟล์ปฏิทินโหราศาสตร์ ManvsDay -> JSON ใน src/lib/bazi/data/almanac/

รัน:  PYTHONUTF8=1 C:/Users/ASUS/miniconda3/python.exe scripts/extract-almanac-tables.py

ที่มา (knownlage/ManvsDay/):
  - ปฏิทิน 2569.xlsx           : ปฏิทินรายวัน (ม.ค.-ก.ค. 2569) -> ตาราง lookup ต่อเสาวัน + score (เสาวัน x เสาเดือน)
  - 2450-2600...สารท....xlsx   : ขอบสารท + เสาปี/เดือน พ.ศ.2450-2600 (fallback / ตรวจ)

โครงสร้างไฟล์ปฏิทิน (ดู memory: manvsday-almanac-decode):
  บล็อก 1 วันเริ่มที่ "แถววันในสัปดาห์" (col A = ชื่อวันไทย).
  แถวถัดมา: col A = เลขวันที่(int), col E = เสาวัน (干@E, 支@E+1),
            col G = เสาเดือนของวันนั้น, col I = เสาปี, col C = day-officer (+desc แถวถัดไป).
  คะแนน 12 คอลัมน์ M..X (กลุ่ม T/D/DM/M/Y), แถวถัดไป = ค่า max.
  แถว header (h): Y/Z = เวลามงคล(โค้ด B/ช่วงเวลา), AA/AB = เทพประจำวัน(คีย์/ชื่อ),
            AC/AD = ธาตุ/สีมงคล, AE = ทิศโชคลาภ, AF = ทิศอสูรวัน,
            AG/AH/AI = เทพอุปถัมป์(คีย์/เลข/คำ), AJ..AY = 8 ประตู(八門) + ทิศ.
  ไฟล์มีบล็อกซ้ำ/คอลัมน์เลื่อน -> dedup เก็บบล็อกแรกต่อ (ชีต,วันที่) + majority vote ต่อเสาวัน.
"""
import json, os, re, sys, unicodedata
from collections import defaultdict, Counter
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "knownlage", "ManvsDay")
OUT = os.path.join(ROOT, "src", "lib", "bazi", "data", "almanac")
os.makedirs(OUT, exist_ok=True)

STEMS = "甲乙丙丁戊己庚辛壬癸"
BRANCHES = "子丑寅卯辰巳午未申酉戌亥"
SS, BR = set(STEMS), set(BRANCHES)
WD = {"อาทิตย์", "จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์"}

# ปรับคำพ้อง/สะกดผิดในไฟล์ต้นฉบับให้เป็นรูปมาตรฐานเดียว
OFFICER_CANON = {
    "ดูแลเอาใส่ใจ": "ดูแลเอาใจใส่",
    "ดูแลใส่ใจ": "ดูแลเอาใจใส่",
    "อำานาจ": "อำนาจ",
    "พระสังกัจจาย์": "พระสังกัจจายน์",
}

def gz_index(stem, branch):
    s, b = STEMS.index(stem), BRANCHES.index(branch)
    for n in range(60):
        if n % 10 == s and n % 12 == b:
            return n
    return -1

def gz_name(n):
    return STEMS[n % 10] + BRANCHES[n % 12]

def _norm_cjk(s):
    # normalize เฉพาะ CJK compatibility ideographs (U+F900-FAFF เช่น 辰=U+F971)
    # ห้าม NFKC ทั้งสตริง เพราะจะทำลายสระไทย (ำ U+0E33 -> ํ+า)
    return "".join(unicodedata.normalize("NFKC", ch) if "豈" <= ch <= "﫿" else ch for ch in s)


def v(ws, r, c):
    val = ws.cell(r, c).value
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    return _norm_cjk(str(val)).strip()

def canon(x):
    return OFFICER_CANON.get(x, x) if isinstance(x, str) else x


def cell_in(ws, r, c, members):
    """membership ที่ normalize ช่องว่าง/CJK-compat glyph (เช่น branch 辰 = U+F971)."""
    val = v(ws, r, c)
    return isinstance(val, str) and val in members


def find_day_blocks(ws):
    """คืน list ของ (header_row, day_row) ต่อบล็อกวัน (เก็บบล็อกแรกต่อเลขวันที่)."""
    headers = [r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value).strip() in WD]
    headers.append(ws.max_row + 1)
    blocks, seen = [], set()
    for i in range(len(headers) - 1):
        h, nxt = headers[i], headers[i + 1]
        day_row = None
        for r in range(h, min(h + 4, nxt)):
            a = ws.cell(r, 1).value
            if isinstance(a, (int, float)) and 1 <= a <= 31 and cell_in(ws, r, 5, SS) and cell_in(ws, r + 1, 5, BR):
                day_row = r
                break
        if day_row is None:
            continue
        dn = int(ws.cell(day_row, 1).value)
        key = (ws.title, dn)
        if key in seen:
            continue
        seen.add(key)
        blocks.append((h, day_row, nxt))
    return blocks


# map ชื่อชีต -> เดือนปฏิทินสากล (ค.ศ. 2026 = พ.ศ. 2569)
SHEET_MONTH = {"jan": 1, "feb": 2, "mar": 3, "apirl": 4, "may": 5, "june": 6, "july": 7}


def sheet_month(title):
    for token, mon in SHEET_MONTH.items():
        if token in title:
            return mon
    return None


def extract_calendar(golden_out=None):
    path = os.path.join(SRC, "ปฏิทิน 2569.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)

    # เกือบทุกชั้น (gates/八神/สี/เวลามงคล/เทพ/ทิศ/อุปถัมป์/คะแนน) เปลี่ยนตามฤดู
    # -> เก็บ "เรคคอร์ดเต็มต่อบล็อก" แล้วโหวตเสียงข้างมากตามคีย์ (เสาวัน × month-branch);
    # day-pillar-table = โหวตรวมทุกเดือน ใช้เป็น fallback เดือน autumn (申酉戌亥)
    blocks = []                                 # (gz, month_branch, date|None, rec)
    month_tab = defaultdict(Counter)            # gz_month -> (field,value) -> count
    year_tab = defaultdict(Counter)
    spirit_legend = defaultdict(Counter)        # 八神 char -> Counter(tuple 4 คีย์เวิร์ด)
    gate_legend = {}                            # 八門 char -> ไทย (จาก AJ1:AY1)
    hour_god_legend = {}                        # B1..B12 -> {god, meaning, score, good} (12 時辰 黃道黑道)
    golden = []
    GOOD_HOUR_CODES = {"B1", "B2", "B5", "B6", "B8"}  # ยามดี 青龍/明堂/金匱/天德/玉堂

    SCORE_COLS = list(range(13, 25))            # M..X
    GATE_COLS = list(range(36, 52, 2))          # AJ,AL,...,AX (8 ประตู/八神)

    def scalar(x):
        return x if x is not None else 0

    for sh in wb.sheetnames:
        ws = wb[sh]
        mon = sheet_month(sh)
        # legend ชื่อประตู 八門 -> ไทย (แถว 1: AJ1=開/AK1=เปิด ...) — เอาจากชีตแรก (ไม่เลื่อน)
        if not gate_legend:
            for c in GATE_COLS:
                g, th = v(ws, 1, c), v(ws, 1, c + 1)
                if g and th:
                    gate_legend[g] = th
        # legend 12 時辰 黃道黑道 (B1..B12) จากตารางท้ายชีต: หาแถวที่ J=='B1' แล้วอ่าน 12 แถว
        if not hour_god_legend:
            for r in range(1, ws.max_row + 1):
                if v(ws, r, 10) == "B1":  # คอลัมน์ J
                    for k in range(12):
                        code = v(ws, r + k, 10)       # J = B-code
                        score = v(ws, r + k, 11)      # K = คะแนน
                        god = v(ws, r + k, 12)        # L = ชื่อเทพยาม
                        meaning = v(ws, r + k, 13)    # M = ความหมาย
                        if code:
                            hour_god_legend[code] = {
                                "god": god, "meaning": meaning, "score": score,
                                "good": code in GOOD_HOUR_CODES,
                            }
                    break
        # ----- ตารางเดือน/ปี จาก header ของชีต -----
        m_pillar = (v(ws, 3, 7), v(ws, 4, 7))   # G3/G4 = เสาเดือนหัวชีต
        y_pillar = (v(ws, 3, 9), v(ws, 4, 9))   # I3/I4 = เสาปี
        if m_pillar[0] in SS and m_pillar[1] in BR:
            gm = gz_name(gz_index(*m_pillar))
            month_tab[gm][("deity_key", v(ws, 3, 27))] += 1
            month_tab[gm][("deity", v(ws, 3, 28))] += 1
            month_tab[gm][("caishen_dir", v(ws, 3, 30))] += 1
            month_tab[gm][("lap_dir", v(ws, 3, 31))] += 1
            month_tab[gm][("asura_dir", v(ws, 3, 32))] += 1
        if y_pillar[0] in SS and y_pillar[1] in BR:
            year_tab[gz_name(gz_index(*y_pillar))][("asura_dir", v(ws, 5, 9))] += 1

        # ----- บล็อกรายวัน -----
        for h, dr, nxt in find_day_blocks(ws):
            stem, branch = v(ws, dr, 5), v(ws, dr + 1, 5)
            if stem not in SS or branch not in BR:
                continue
            mp = (v(ws, dr, 7), v(ws, dr + 1, 7))
            if mp[1] not in BR:
                continue
            gz, mb = gz_name(gz_index(stem, branch)), mp[1]
            # header row (แถววันในสัปดาห์): เทพ/สี1/ทิศ/อุปถัมป์1/ประตู
            # ไม่ fix คอลัมน์: บางชีต (เช่น july) แทรก 2 คอลัมน์คะแนนพิเศษ -> บล็อกขวาเลื่อน +off
            # หา off จากตำแหน่งโค้ดยาม B\d+ (ปกติอยู่คอลัมน์ 25)
            off = 0
            for hrow in (h, h + 1):
                for c in range(25, 30):
                    val = v(ws, hrow, c)
                    if isinstance(val, str) and re.fullmatch(r"B\d+", val):
                        off = c - 25
                        break
                else:
                    continue
                break
            DC, AA, AB = 25 + off, 27 + off, 28 + off  # เวลามงคล / deity key / deity
            ACc, ADc = 29 + off, 30 + off              # สี ธาตุ/ข้อความ
            AEc, AFc = 31 + off, 32 + off              # ทิศโชคลาภ / ทิศอสูร
            AGc, AHc, AIc = 33 + off, 34 + off, 35 + off  # อุปถัมป์
            GATE0 = 36 + off
            gate_cols = list(range(GATE0, GATE0 + 16, 2))
            hr = h if (v(ws, h, AB) and v(ws, h, AA) in (SS | BR)) else h + 1

            # holy-day marker: คอลัมน์ C ในบล็อก == "วันพระจีน"/"วันพระ"
            holy = any(isinstance(v(ws, r, 3), str) and "วันพระ" in v(ws, r, 3) for r in range(h, nxt))

            # เทพประจำวัน อาจมี 2 องค์ (header row + dayrow) เก็บเป็น list
            deities = []
            for r in (hr, hr + 1):
                nm = canon(v(ws, r, AB))
                if nm and v(ws, r, AA) in (SS | BR) and nm not in deities:
                    deities.append(nm)

            rec = {
                "officer": canon(v(ws, dr, 3)),
                "officer_desc": v(ws, dr + 1, 3),
                "deity_key": canon(v(ws, hr, AA)),
                "deity": canon(v(ws, hr, AB)),
                "deities": deities,
                "color_primary": [v(ws, hr, ACc), v(ws, hr, ADc)],
                "color_secondary": [v(ws, dr, ACc), v(ws, dr, ADc)],
                "lucky_dir": v(ws, hr, AEc),
                "asura_dir": v(ws, hr, AFc),
                "patrons": [
                    [v(ws, hr, AGc), v(ws, hr, AHc), v(ws, hr, AIc)],  # อุปถัมป์ 1 (header)
                    [v(ws, dr, AGc), v(ws, dr, AHc), v(ws, dr, AIc)],  # อุปถัมป์ 2 (dayrow)
                ],
                "gates": [[v(ws, hr, c), v(ws, hr, c + 1)] for c in gate_cols],   # 八門 name+dir
                "spirits": [v(ws, dr, c) for c in gate_cols],                     # 八神 arrangement
                "lucky_hours": [[v(ws, r, DC), v(ws, r, DC + 1)]
                                for r in range(h, nxt) if v(ws, r, DC) and v(ws, r, DC + 1)],
                "scores": [scalar(v(ws, dr, c)) for c in SCORE_COLS],
                "max": [scalar(v(ws, dr + 1, c)) for c in SCORE_COLS],
                "holy_day": holy,
                # เสาเดือน/ปี เต็มจากต้นฉบับ (ไว้ validate source-table; ไม่เข้า DISPLAY_FIELDS)
                "month_pillar": (mp[0] + mp[1]),
                "year_pillar": ((v(ws, dr, 9) or "") + (v(ws, dr + 1, 9) or "")),
            }
            blocks.append((gz, mb, f"2026-{mon:02d}-{int(ws.cell(dr, 1).value):02d}" if mon else None, rec))

            # legend 八神 -> 4 คีย์เวิร์ด (คอลัมน์ขวาของเทพ แถว dr..dr+3)
            for c in gate_cols:
                sp = v(ws, dr, c)
                kws = tuple(v(ws, dr + k, c + 1) for k in range(4))
                if sp and all(kws):
                    spirit_legend[sp][kws] += 1

    # ---- โหวต record ----
    DISPLAY_FIELDS = ["officer", "officer_desc", "deity_key", "deity", "deities", "color_primary",
                      "color_secondary", "lucky_dir", "asura_dir", "patrons", "gates",
                      "spirits", "lucky_hours", "scores", "max", "holy_day"]

    def vote_recs(recs, field):
        c = Counter(json.dumps(r[field], ensure_ascii=False, sort_keys=True) for r in recs)
        return json.loads(c.most_common(1)[0][0])

    by_dm, by_d = defaultdict(list), defaultdict(list)
    for gz, mb, _date, rec in blocks:
        by_dm[(gz, mb)].append(rec)
        by_d[gz].append(rec)

    day_month_table = {}
    for (gz, mb), recs in by_dm.items():
        entry = {f: vote_recs(recs, f) for f in DISPLAY_FIELDS}
        entry["day"], entry["month_branch"] = gz, mb
        day_month_table[f"{gz}|{mb}"] = entry

    day_table = {}
    for n in range(60):
        gz = gz_name(n)
        recs = by_d.get(gz, [])
        entry = {"index": n, "pillar": gz}
        for f in DISPLAY_FIELDS:
            entry[f] = vote_recs(recs, f) if recs else None
        day_table[gz] = entry

    # score table (คงไว้เพื่อ buildStrength) — ดึงจาก day-month record
    score_table = {
        key: {"day": e["day"], "month_branch": e["month_branch"], "scores": e["scores"], "max": e["max"]}
        for key, e in day_month_table.items()
    }

    spirit_out = {sp: list(cnt.most_common(1)[0][0]) for sp, cnt in spirit_legend.items()}

    month_out = {gz: {k: (Counter({kk[1]: c for kk, c in fields.items() if kk[0] == k}).most_common(1)[0][0]
                          if any(kk[0] == k for kk in fields) else None)
                      for k in ("deity_key", "deity", "caishen_dir", "lap_dir", "asura_dir")}
                 for gz, fields in month_tab.items()}
    year_out = {gz: {"asura_dir": Counter({kk[1]: c for kk, c in fields.items() if kk[0] == "asura_dir"}).most_common(1)[0][0]}
                for gz, fields in year_tab.items()}

    if golden_out is not None:
        seen_dates, uniq = set(), []
        for gz, mb, date, rec in blocks:
            if date is None or date in seen_dates:
                continue
            seen_dates.add(date)
            uniq.append({"date": date, "day_pillar": gz, "month_branch": mb, **rec})
        uniq.sort(key=lambda g: g["date"])
        golden_out.extend(uniq)

    return (day_table, score_table, day_month_table, month_out, year_out,
            spirit_out, gate_legend, hour_god_legend)


def extract_solar_terms():
    """ขอบสารทต่อ (ปีพ.ศ., เดือน) จากไฟล์ 2450-2600 -> {be: {month: {...}}}."""
    path = os.path.join(SRC, "2450-2600วันเปลี่ยนสารทเล็กสารทใหญ่.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    out = {}
    for r in range(2, ws.max_row + 1):
        be = ws.cell(r, 2).value
        mon = ws.cell(r, 5).value
        if be is None or mon is None:
            continue
        try:
            be = int(be); mon = int(mon)
        except (TypeError, ValueError):
            continue
        ystem, ybr = v(ws, r, 3), v(ws, r, 4)
        mstem, mbr = v(ws, r, 7), v(ws, r, 8)
        def t(c):
            val = ws.cell(r, c).value
            return val.strftime("%H:%M") if hasattr(val, "strftime") else (str(val) if val is not None else None)
        out.setdefault(str(be), {})[str(mon)] = {
            "year_pillar": (ystem or "") + (ybr or ""),
            "month_pillar": (mstem or "") + (mbr or ""),
            "start_day": v(ws, r, 9), "start_month": v(ws, r, 10), "start_time": t(11),
            "big_start_day": v(ws, r, 15), "big_start_month": v(ws, r, 16), "big_start_time": t(17),
        }
    return out


def main():
    golden = []
    (day_table, score_table, day_month_table, month_out, year_out,
     spirit_legend, gate_legend, hour_god_legend) = extract_calendar(golden_out=golden)
    solar = extract_solar_terms()

    def dump(path, name, obj):
        p = os.path.join(path, name)
        os.makedirs(path, exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
        print(f"  wrote {name}: {len(obj)} entries")

    print("Extracted:")
    dump(OUT, "day-pillar-table.json", day_table)
    dump(OUT, "day-month-table.json", day_month_table)
    dump(OUT, "score-day-month.json", score_table)
    dump(OUT, "month-pillar-table.json", month_out)
    dump(OUT, "year-pillar-table.json", year_out)
    dump(OUT, "spirit-legend.json", spirit_legend)
    dump(OUT, "gate-legend.json", gate_legend)
    dump(OUT, "hour-god-legend.json", hour_god_legend)
    dump(OUT, "solar-terms-2450-2600.json", solar)
    dump(os.path.join(ROOT, "tests", "fixtures"), "almanac-2569-golden.json", golden)

    # ---- diagnostics ----
    covered_day = sum(1 for r in day_table.values() if r["officer"])
    month_branches = sorted({k.split("|")[1] for k in score_table}, key=lambda b: BRANCHES.index(b))
    print(f"\nDiagnostics:")
    print(f"  day-pillars with officer: {covered_day}/60")
    print(f"  score (day,month-branch) pairs: {len(score_table)}")
    print(f"  month-branches covered by scores: {''.join(month_branches)}  (missing: "
          f"{''.join(b for b in BRANCHES if b not in month_branches)})")
    print(f"  solar-term years: {len(solar)} (range {min(solar)}-{max(solar)})")


if __name__ == "__main__":
    main()
