# -*- coding: utf-8 -*-
"""
Distill the 4 "ค่าพลัง %ของดวง" Excel files in knownlage/การหาค่าพลัง/ into JSON
assets consumed by the domain-power engine (src/lib/bazi/symbolic-engine.domain-power.ts).

Key finding: the 60x60 (3600-entry) pair-coefficient matrix is IDENTICAL across the
wealth / work / learning files (verified by diff). Each domain only differs in which
pair of chart pillars it feeds into that one shared matrix:
  - career   : matrix[dayPillar | monthPillar]
  - learning : matrix[dayPillar | hourPillar]
  - wealth   : matrix over the chart's 财 (wealth-star) positions vs day & month, averaged
  - friends  : a separate per-day-pillar table (file: ...เพื่อนฝูง...) with interp text

Outputs (src/lib/bazi/data/domain-power/):
  matrix.json   - { "<dayStem><dayBranch>|<otherStem><otherBranch>": coefficient(0..1) }
  friends.json  - { "<dayStem><dayBranch>": { coefficient, interpretation } }
  cases.json    - validation fixtures:
                    { careerLearningPairs:[{day,other,coefficient}],
                      wealth:[{label,pillars{hour,day,month,year},expectedScore,rows}] }

Run:  python scripts/parse-domain-power-xlsx.py
"""
import json
import os
import unicodedata

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC_DIR = os.path.join(ROOT, "knownlage", "การหาค่าพลัง")
OUT_DIR = os.path.join(ROOT, "src", "lib", "bazi", "data", "domain-power")

STEMS = set("甲乙丙丁戊己庚辛壬癸")
BRANCHES = set("子丑寅卯辰巳午未申酉戌亥")

WEALTH_FILE = "ค่าพลังการเงิน %ของดวง(สกิลเรียกทรัพย์).xlsx"
WORK_FILE = "ค่าพลังการทำงาน_การศึกษา %ของดวง(สกิลตัวท็อป).xlsx"
LEARN_FILE = "ค่าพลังการความเข้าใจ %ของดวง(สกิลเรียนรู้).xlsx"
FRIENDS_FILE = "ค่าพลังการเพื่อนฝูง %ของดวง(พลังเพื่อน).xlsx"


def norm(value):
    """Trim + NFKC-canonicalise (fixes U+F971 -> 辰 compatibility-ideograph bug)."""
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load(fname):
    return openpyxl.load_workbook(os.path.join(SRC_DIR, fname), data_only=True)


# ---------------------------------------------------------------------------
# 1. matrix.json  (shared 60x60 pair-coefficient matrix)
# ---------------------------------------------------------------------------
def parse_matrix(wb):
    """Pair sheets named "<elem><polarity>,<branchNo>" (e.g. ม1,1=甲子). Each is one
    day-pillar x 60 other-pillars in 2-row blocks:
      top:    B=ourStem D=otherStem ... H=coefficient(0..1) I=H*300 J,K,L=components
      bottom: B=ourBranch D=otherBranch
    coefficient (H) is the value the วิธีคำนวน worksheet calls G in its EX blocks."""
    out = {}
    for ws in wb.worksheets:
        title = ws.title.strip()
        if "," not in title or title[0] not in "มฟดทน":
            continue
        rows = list(ws.iter_rows(values_only=True))

        def cell(r, idx):
            if r < 0 or r >= len(rows):
                return ""
            row = rows[r]
            return norm(row[idx]) if idx < len(row) else ""

        i = 0
        while i < len(rows):
            our_stem, other_stem = cell(i, 1), cell(i, 3)
            if our_stem in STEMS and other_stem in STEMS:
                our_branch, other_branch = cell(i + 1, 1), cell(i + 1, 3)
                if our_branch in BRANCHES and other_branch in BRANCHES:
                    coeff = num(rows[i][7]) if len(rows[i]) > 7 else None
                    key = f"{our_stem}{our_branch}|{other_stem}{other_branch}"
                    out[key] = round(coeff, 6) if coeff is not None else None
                    i += 2
                    continue
            i += 1
    return out


# ---------------------------------------------------------------------------
# 2. friends.json  (per day-pillar coefficient + interpretation)
# ---------------------------------------------------------------------------
def parse_friends(wb):
    """Single sheet "ค่าพลังเพื่อน": per day master, 6 two-row blocks (one per branch).
      top:    A=dayCode B=dayStem C/D=stageCodes E=coefficient F=E*200 ... J=interpretation
      bottom: A=branchNo B=dayBranch C/D=stageNames
    Keyed by full day pillar (e.g. 甲子)."""
    ws = wb["ค่าพลังเพื่อน"]
    rows = list(ws.iter_rows(values_only=True))

    def cell(r, idx):
        if r < 0 or r >= len(rows):
            return ""
        row = rows[r]
        return norm(row[idx]) if idx < len(row) else ""

    out = {}
    i = 0
    while i < len(rows):
        day_stem = cell(i, 1)          # B top
        day_branch = cell(i + 1, 1)    # B bottom
        if day_stem in STEMS and day_branch in BRANCHES:
            coeff = num(rows[i][4]) if len(rows[i]) > 4 else None   # E
            interp = cell(i, 9)                                     # J
            key = f"{day_stem}{day_branch}"
            out[key] = {
                "coefficient": round(coeff, 6) if coeff is not None else None,
                "interpretation": interp or None,
            }
            i += 2
            continue
        i += 1
    return out


# ---------------------------------------------------------------------------
# 3. cases.json  (validation fixtures)
# ---------------------------------------------------------------------------
def parse_career_learning_pairs(wb, sheet_name):
    """EX blocks in the วิธีคำนวน sheet:
      row n:   A=EX B=<idx> C=<dayStem>  D=Vs E=<otherStem> F== G=<coefficient>
      row n+1:         C=<dayBranch>          E=<otherBranch>
    Yields {day, other, coefficient}."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    def cell(r, idx):
        if r < 0 or r >= len(rows):
            return ""
        row = rows[r]
        return norm(row[idx]) if idx < len(row) else ""

    cases = []
    for i, _ in enumerate(rows):
        if cell(i, 0) != "EX":
            continue
        day_stem, other_stem = cell(i, 2), cell(i, 4)
        day_branch, other_branch = cell(i + 1, 2), cell(i + 1, 4)
        coeff = num(rows[i][6]) if len(rows[i]) > 6 else None
        if day_stem in STEMS and other_stem in STEMS and day_branch in BRANCHES \
                and other_branch in BRANCHES and coeff is not None:
            cases.append({
                "day": f"{day_stem}{day_branch}",
                "other": f"{other_stem}{other_branch}",
                "coefficient": round(coeff, 6),
            })
    return cases


def parse_wealth_cases(wb):
    """The วิธีคำนวนพลังการเงิน sheet holds real birth-chart worked examples:
      header row: B='<thai date> เวลา <hh:mm> เพศ<ชาย|หญิง>'
      next 2 rows: B/C/D/E = (hour,day,month,year) stems then branches
      a 'ค่าพลังการเงินในดวง' row: N=avg-coefficient O=sum P=row-count
    Captures the inputs + the worksheet's own final score for end-to-end validation."""
    ws = wb["วิธีคำนวนพลังการเงิน"]

    def c(r, col):
        return norm(ws.cell(r, col).value)

    def n(r, col):
        return num(ws.cell(r, col).value)

    cases = []
    r = 1
    while r <= ws.max_row:
        label = c(r, 2)
        if ("เวลา" in label or "เพศ" in label) and any(ch.isdigit() for ch in label):
            # pillars on the two rows that follow once we hit a ยาม/วัน/เดือน/ปี header
            hdr = None
            for rr in range(r + 1, min(r + 4, ws.max_row + 1)):
                if c(rr, 2) == "ยาม" and c(rr, 3) == "วัน":
                    hdr = rr
                    break
            pillars = None
            if hdr:
                stems = [c(hdr + 1, col) for col in range(2, 6)]
                branches = [c(hdr + 2, col) for col in range(2, 6)]
                if all(s in STEMS for s in stems) and all(b in BRANCHES for b in branches):
                    pillars = {
                        "hour": stems[0] + branches[0],
                        "day": stems[1] + branches[1],
                        "month": stems[2] + branches[2],
                        "year": stems[3] + branches[3],
                    }
            # find this example's final score row + the ground-truth pairings it used,
            # scanning until the next example header
            final = None
            pairs = []   # (basePillar, laphPillar) compared via the matrix
            selfs = []   # day-pillar-sitting-on-wealth rows
            for rr in range(r + 1, ws.max_row + 1):
                nxt = c(rr, 2)
                if rr != r and ("เวลา" in nxt or "เพศ" in nxt) and any(ch.isdigit() for ch in nxt):
                    break
                if c(rr, 8) == "ค่าพลังการเงินในดวง":
                    final = {"avg": n(rr, 14), "sum": n(rr, 15), "count": n(rr, 16)}
                    break
                # detail pair row: O(15)=coefficient, I(9)/K(11)=stems (top), branches on next row
                o = n(rr, 15)
                i_s, k_s = c(rr, 9), c(rr, 11)
                i_b, k_b = c(rr + 1, 9), c(rr + 1, 11)
                if o is not None and 0 < o < 1 and i_s in STEMS and k_s in STEMS \
                        and i_b in BRANCHES and k_b in BRANCHES:
                    pairs.append({"base": i_s + i_b, "laph": k_s + k_b, "coefficient": round(o, 6)})
                # self/day-sitting row: L(12)=coefficient, J(10)=code 'Axx'
                l = n(rr, 12)
                if l is not None and 0 < l < 1 and i_s in STEMS and i_b in BRANCHES \
                        and c(rr, 10).startswith("A"):
                    selfs.append({"pillar": i_s + i_b, "coefficient": round(l, 6)})
            if pillars and final and final["avg"] is not None:
                cases.append({
                    "label": label,
                    "pillars": pillars,
                    "dayMaster": pillars["day"][0],
                    "expectedAvg": round(final["avg"], 6),
                    "expectedScore": round(final["avg"] * 100, 2),
                    "rowCount": final["count"],
                    "pairs": pairs,
                    "selfRows": selfs,
                })
        r += 1
    return cases


# ---------------------------------------------------------------------------
def write_json(name, data):
    path = os.path.join(OUT_DIR, name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wealth_wb = load(WEALTH_FILE)
    work_wb = load(WORK_FILE)
    learn_wb = load(LEARN_FILE)
    friends_wb = load(FRIENDS_FILE)

    # ---- matrix (verify identical across the 3 files) ----
    matrix = parse_matrix(wealth_wb)
    for tag, wb in (("work", work_wb), ("learn", learn_wb)):
        other = parse_matrix(wb)
        diffs = sum(1 for k, v in matrix.items() if other.get(k) != v)
        missing = sum(1 for k in matrix if k not in other)
        assert diffs == 0 and missing == 0 and len(other) == len(matrix), \
            f"matrix mismatch vs {tag}: diffs={diffs} missing={missing} " \
            f"sizes={len(matrix)}/{len(other)}"
    write_json("matrix.json", matrix)
    print(f"matrix.json: {len(matrix)} pairs (identical across wealth/work/learn)")

    # ---- friends ----
    friends = parse_friends(friends_wb)
    write_json("friends.json", friends)
    with_text = sum(1 for v in friends.values() if v["interpretation"])
    print(f"friends.json: {len(friends)} day-pillars ({with_text} with interpretation)")

    # ---- cases ----
    career_pairs = parse_career_learning_pairs(work_wb, work_wb.sheetnames[0])
    learn_pairs = parse_career_learning_pairs(learn_wb, learn_wb.sheetnames[0])
    # EX pairs validate the matrix lookup directly
    pairs = career_pairs + learn_pairs
    bad = [p for p in pairs if matrix.get(f"{p['day']}|{p['other']}") != p["coefficient"]]
    print(f"cases: career/learning EX pairs={len(pairs)} matrix-mismatch={len(bad)}")
    for p in bad:
        print("  MISMATCH", p, "matrix=", matrix.get(f"{p['day']}|{p['other']}"))
    wealth_cases = parse_wealth_cases(wealth_wb)
    write_json("cases.json", {
        "careerLearningPairs": pairs,
        "wealth": wealth_cases,
    })
    print(f"cases.json: {len(pairs)} pair-cases, {len(wealth_cases)} wealth birth-cases")


if __name__ == "__main__":
    main()
