#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
สกัด "ทิศ 八神 ระดับปี/เดือน" จากหัวชีตปฏิทิน 2569 -> เติม field spirit_dirs
ลง year-pillar-table.json (ตามเสาปี) และ month-pillar-table.json (ตามเสาเดือน)
โดย "merge" ไม่ทับ field เดิม

รัน:  PYTHONUTF8=1 C:/Users/ASUS/miniconda3/python.exe scripts/extract-spirit-dirs.py

โครงสร้าง (ตรวจจากไฟล์): หัวชีต cols 36-51
  r3 = ทิศ 八神 ระดับปี  (jan=เสาปี 乙巳, feb-jul=丙午)  -> [char, dir] x8
  r4 = ทิศ 八神 ระดับเดือน (ตามเสาเดือนของชีต)          -> [char, dir] x8
"""
import json, os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "knownlage", "ManvsDay", "ปฏิทิน 2569.xlsx")
OUT = os.path.join(ROOT, "src", "lib", "bazi", "data", "almanac")
DIRS = {"N", "S", "E", "W", "NE", "NW", "SE", "SW"}
SPIRITS = set("天地玄虎合陰蛇符陳雀")


def grab(ws, r):
    """อ่าน [char, dir] 8 คู่จากแถว r (cols 36..51)"""
    out = []
    c = 36
    while c <= 51:
        ch = ws.cell(r, c).value
        dr = ws.cell(r, c + 1).value
        if ch in SPIRITS and dr in DIRS:
            out.append([ch, dr])
        c += 2
    return out


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    year_dirs = {}   # เสาปี -> [[char,dir]]
    month_dirs = {}  # เสาเดือน -> [[char,dir]]
    for sh in wb.sheetnames:
        ws = wb[sh]
        month_gz = sh.split("-")[-1][-2:]  # เช่น "甲午"
        # jan (己丑) อยู่ก่อนลิบชุน -> เสาปี 乙巳; ที่เหลือ 丙午
        year_gz = "乙巳" if month_gz == "己丑" else "丙午"
        r3, r4 = grab(ws, 3), grab(ws, 4)
        if r3 and year_gz not in year_dirs:
            year_dirs[year_gz] = r3
        if r4:
            month_dirs[month_gz] = r4

    for fname, add in [("year-pillar-table.json", year_dirs), ("month-pillar-table.json", month_dirs)]:
        path = os.path.join(OUT, fname)
        data = json.load(open(path, encoding="utf-8"))
        for gz, dirs in add.items():
            if gz not in data:
                data[gz] = {}
            data[gz]["spirit_dirs"] = dirs
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.write("\n")
        print(f"updated {fname}: +spirit_dirs for {list(add.keys())}")


if __name__ == "__main__":
    main()
